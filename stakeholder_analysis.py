# -*- coding: utf-8 -*-

import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import datetime
import matplotlib.dates as mdates
from stakeholder_analysis_util import stakeholder_analysis 

from openpyxl import Workbook

plt.rcParams['figure.figsize'] = [11, 5]

data_path = r'Data\fishery_lemmas_sentences_labeled.csv'
pers_path = r'Data\common_pers_cleaned.xlsx'
orgs_path = r'Data\common_org_cleaned.xlsx'
quota_path = r'Data\data, fish & fisheries, SD22-24.xlsx'
start_year = 2009
end_year = 2022

def load_and_preprocess_data(data_path, pers_path, orgs_path, quota_path, newspapers = 'all'):
    # Specify data types for columns
    dtype_dict = {
        'Unnamed: 0.1': 'int64',
        'Unnamed: 0': 'int64',
        'id': 'str',
        'year': 'int64',
        'Journal': 'str',
        'type of newspaper (regional, national)': 'str',
        'Date': 'str',  
        'Category': 'str',
        'preheading': 'str',
        'Klima': 'float64',
        'Naturschutz': 'float64',
        'Fischerei': 'float64',
        'Unnamed: 23': 'str',
        'lemmas': 'str',
        'text': 'str',
        'word_count': 'int64',
        'Label': 'int64'
    }

    # Load data with specified dtypes
    data = pd.read_csv(data_path, dtype=dtype_dict, low_memory=False)
    data["Date"] = pd.to_datetime(data["Date"], format='%Y-%m-%d')
    data = data.rename(columns={'Lemmas': 'lemmas'})
    data = data.dropna(subset=['lemmas'])
    
    if newspapers == 'regional':
        data = data[data['type of newspaper (regional, national)'] == 'regional']
        
    elif newspapers == 'national':
        data = data[data['type of newspaper (regional, national)'] == 'national']
        
    # Load and transform lists with organisations and names
    common_pers = pd.read_excel(pers_path)
    common_orgs = pd.read_excel(orgs_path)
    
    common_orgs.dropna(subset=['Name'], inplace=True)
    common_orgs['Lemma1'] = [common_orgs['Name'].iloc[idx].lower() if name != name else name for idx, name in enumerate(common_orgs['Lemma1'])]
    common_orgs['Abkürzung'] = [common_orgs['Abkürzung'].iloc[idx].lower() if name == name else float('nan') for idx, name in enumerate(common_orgs['Abkürzung'])]
    
    quota_dates = pd.read_excel(quota_path, sheet_name='dates, advice - quota ')['quota decision']
    quota_dates = quota_dates[(quota_dates.dt.year >= start_year) & (quota_dates.dt.year <= end_year)][:-1].reset_index(drop=True)
    
    common_pers['Name'] = [nam.lower() for nam in common_pers['Name']]
    common_pers['Name'] = common_pers['Name'].str.replace("von ", "", regex=False)
    
    
    # Select all stakeholders
    stakeholders_pers_list = list(common_pers[common_pers['Stakeholder-Gruppe'].str.contains('politics|management|science|fisheries|engo|rf', na=False)]['Name'])
    stakeholders_orgs_list = list(common_orgs[common_orgs['Stakeholder-Gruppe'].str.contains('politics|management|science|fisheries|engo|rf', na=False)]['Name'])

    return data, common_pers, common_orgs, quota_dates, stakeholders_pers_list, stakeholders_orgs_list

def run_stakeholder_analysis(data, common_pers, common_orgs, stakeholders_pers_list, stakeholders_orgs_list, start_year, end_year, quota_dates):
    return stakeholder_analysis(
        data, 
        common_pers, 
        common_orgs, 
        stakeholders_pers_list, 
        stakeholders_orgs_list, 
        start_year, 
        end_year, 
        quota_dates
    )

def analyze_by_group(common_pers, common_orgs, group_name, data, start_year, end_year, quota_dates):
    stakeholders_pers_list = list(common_pers[common_pers['Stakeholder-Gruppe'].str.contains(group_name, na=False)]['Name'])
    stakeholders_orgs_list = list(common_orgs[common_orgs['Stakeholder-Gruppe'].str.contains(group_name, na=False)]['Name'])
    return run_stakeholder_analysis(data, common_pers, common_orgs, stakeholders_pers_list, stakeholders_orgs_list, start_year, end_year, quota_dates)

def calculate_summary_stats(data):
    return data.describe()[['mean', 'std', 'min', 'max']]

def combine_politics_management(df):
    # Combine 'politics' and 'management' into one group
    df['Stakeholder-Gruppe'] = df['Stakeholder-Gruppe'].replace(
        ['politics', 'management'], 'politics|management'
    )
    return df

def write_summary_to_excel(summary_stats, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary Statistics"
    
    # Write header for Table S6 (Sentiment by groups)
    ws.append(["Table S6. Stakeholder sentiment by groups"])
    ws.append([])
    ws.append(["Individuals"])
    ws.append(["Yearly sentiment politics & authorities individuals", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_persons_sentiment'].round(3)
        ws.append([f"Yearly sentiment {group.replace('|', ' & ')} individuals"] + list(stats.values))
    
    ws.append([])
    ws.append(["Organizations"])
    ws.append(["Yearly sentiment Politics & Authorities Organizations", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_organizations_sentiment'].round(3)
        ws.append([f"Yearly sentiment {group.replace('|', ' & ')} organizations"] + list(stats.values))
    
    ws.append([])
    ws.append(["Both"])
    ws.append(["Yearly sentiment Politics & Authorities Both", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_both_sentiment'].round(3)
        ws.append([f"Yearly sentiment {group.replace('|', ' & ')} both"] + list(stats.values))

    # Write header for Table S7 (Shares by groups)
    ws.append([])
    ws.append(["Table S7. Stakeholder shares by groups (multiple occurrences of same stakeholder per articles)"])
    ws.append([])
    ws.append(["Individuals"])
    ws.append(["Yearly share politics & authorities individuals", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_persons_share'].round(3)
        ws.append([f"Yearly share {group.replace('|', ' & ')} individuals"] + list(stats.values))
    
    ws.append([])
    ws.append(["Organizations"])
    ws.append(["Yearly share Politics & Authorities Organizations", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_organizations_share'].round(3)
        ws.append([f"Yearly share {group.replace('|', ' & ')} organizations"] + list(stats.values))
    
    ws.append([])
    ws.append(["Both"])
    ws.append(["Yearly share Politics & Authorities Both", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_both_share'].round(3)
        ws.append([f"Yearly share {group.replace('|', ' & ')} both"] + list(stats.values))

    # Write header for Table S8 (Shares by groups, counting a stakeholder only once per article)
    ws.append([])
    ws.append(["Table S10. Stakeholder shares by groups (counting a stakeholder only once per article)"])
    ws.append([])
    ws.append(["Individuals"])
    ws.append(["Yearly share politics & authorities individuals", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_persons_share_once'].round(3)
        ws.append([f"Yearly share {group.replace('|', ' & ')} individuals"] + list(stats.values))
    
    ws.append([])
    ws.append(["Organizations"])
    ws.append(["Yearly share Politics & Authorities Organizations", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_organizations_share_once'].round(3)
        ws.append([f"Yearly share {group.replace('|', ' & ')} organizations"] + list(stats.values))
    
    ws.append([])
    ws.append(["Both"])
    ws.append(["Yearly share Politics & Authorities Both", "Mean", "S.D.", "Min", "Max"])
    for group in ['politics|management', 'science', 'fisheries', 'engo']:
        stats = summary_stats[f'{group}_both_share_once'].round(3)
        ws.append([f"Yearly share {group.replace('|', ' & ')} both"] + list(stats.values))

    wb.save(filename)

def create_summary_excel_2(summary_stats, filename):
    wb = Workbook()
    
    # Define categories
    groups = ['all', 'politics|management', 'fisheries', 'science', 'engo']
    types = ['persons', 'organizations', 'both']
    measures = ['sentiment', 'share_once', 'share']

    for measure in measures:
        ws = wb.create_sheet(title=f"{measure.capitalize()} Stats")  # Create a new worksheet for each measure

        # Write the header
        ws.append(["Type"] + groups)  # Include 'Type' directly in the header

        # Helper function to add mean values to the Excel
        def add_means(type_):
            row = [type_]  # Starts row with the type
            for group in groups:
                key = f'{group}_{type_}_{measure}'
                mean_value = summary_stats.get(key, {'mean': 'N/A'})['mean']
                row.append(mean_value.round(3))
            ws.append(row)

        # Append data for each type
        for type_ in types:
            add_means(type_)

    # Remove the default sheet created when initiating the workbook
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Save the workbook
    wb.save(filename)
    
def compute_and_format_subtotals(data_frame):
    # Calculate subtotals for each combination of Stakeholder-Gruppe
    subtotals_by_level = data_frame.groupby(['Stakeholder-Gruppe', 'Type', 'Governance level']).sum().reset_index()
    
    return subtotals_by_level
    
def main(data_path, pers_path, orgs_path, quota_path, start_year, end_year):
    data, common_pers, common_orgs, quota_dates, stakeholders_pers_list, stakeholders_orgs_list = load_and_preprocess_data(data_path, pers_path, orgs_path, quota_path)
    
    # Apply the function to both person and organization dataframes
    common_pers = combine_politics_management(common_pers)
    common_orgs = combine_politics_management(common_orgs)
    
    # Initial stakeholder analysis
    results = run_stakeholder_analysis(data, common_pers, common_orgs, stakeholders_pers_list, stakeholders_orgs_list, start_year, end_year, quota_dates)
    stakeholders_pers_sum_all_persons, stakeholders_orgs_sum_all_orgs = results[9], results[10]
    
    # Merge the filtered results with common_pers to get governance level and stakeholder group
    common_pers = common_pers.merge(stakeholders_pers_sum_all_persons, on='Name')
    common_orgs = common_orgs.merge(stakeholders_orgs_sum_all_orgs, on='Name')
    
    # Filtering to exclude 'rf' from the 'Stakeholder' column in common_pers_filtered
    common_pers = common_pers[~common_pers['Stakeholder'].isin(['rf'])]

    # Filtering to exclude 'rf' from the 'Governance level' column in common_orgs_filtered
    common_orgs = common_orgs[~common_orgs['Stakeholder-Gruppe'].isin(['rf'])]

    # Prepare the final table for organizations with the desired format
    orgs_table = common_orgs[['Name', 'Stakeholder-Gruppe', 'Governance level', 'Value']].fillna('')
    orgs_table = orgs_table.sort_values(by=['Stakeholder-Gruppe', 'Governance level', 'Value'], ascending=[True, True, False])
    
    # Prepare the final table for persons with the desired format
    pers_table = common_pers[['Name', 'Stakeholder-Gruppe', 'Governance level', 'Value']].fillna('')
    pers_table = pers_table.sort_values(by=['Stakeholder-Gruppe', 'Governance level', 'Value'], ascending=[True, True, False])
    
    pers_table = pers_table[pers_table['Value'] >= 5].copy()
    orgs_table = orgs_table[orgs_table['Value'] >= 5].copy()
    
    # Create a formatted Excel file
    output_path = 'Results\stakeholder_analysis_count_tabS4a_S4b.xlsx'
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        workbook = writer.book
    
        # Formatting
        bold = workbook.add_format({'bold': True})
        italic = workbook.add_format({'italic': True})
    
        # Writing organizations
        worksheet = workbook.add_worksheet('Organizations')
        worksheet.write('A1', 'Name', bold)
        worksheet.write('B1', 'Value', bold)
        worksheet.write('C1', 'Stakeholder Group', bold)
    
        start_row = 1
        for stakeholder_group in orgs_table['Stakeholder-Gruppe'].unique():
            worksheet.write(start_row, 3, stakeholder_group, italic)
            start_row += 1
            
            for gov_level in orgs_table['Governance level'].unique():
                worksheet.write(start_row, 3, gov_level, italic)
                start_row += 1
                
                for _, row in orgs_table[(orgs_table['Stakeholder-Gruppe'] == stakeholder_group) & (orgs_table['Governance level'] == gov_level)].iterrows():
                    worksheet.write(start_row, 0, row['Governance level'])
                    worksheet.write(start_row, 1, row['Name'])
                    worksheet.write(start_row, 2, row['Value'])
                    start_row += 1
                start_row += 1
            
            start_row += 1
    
        # Writing persons
        worksheet = workbook.add_worksheet('Persons')
        worksheet.write('A1', 'Name', bold)
        worksheet.write('B1', 'Value', bold)
        worksheet.write('C1', 'Stakeholder Group', bold)
    
        start_row = 1
        for stakeholder_group in pers_table['Stakeholder-Gruppe'].unique():
            worksheet.write(start_row, 3, stakeholder_group, italic)
            start_row += 1
            
            for gov_level in orgs_table['Governance level'].unique():
                worksheet.write(start_row, 3, gov_level, italic)
                start_row += 1
            
                for _, row in pers_table[(pers_table['Stakeholder-Gruppe'] == stakeholder_group) & (pers_table['Governance level'] == gov_level)].iterrows():
                    worksheet.write(start_row, 0, row['Governance level'])
                    worksheet.write(start_row, 1, row['Name'])
                    worksheet.write(start_row, 2, row['Value'])
                    start_row += 1
                start_row += 1
            
            start_row += 1
            
    pers_table = pers_table[pers_table['Value'] >= 20].copy()
    orgs_table = orgs_table[orgs_table['Value'] >= 20].copy()
    
    common_pers = common_pers[common_pers['Name'].isin(pers_table['Name'])]
    common_orgs = common_orgs[common_orgs['Name'].isin(orgs_table['Name'])]
    
    results = run_stakeholder_analysis(data, common_pers, common_orgs, stakeholders_pers_list, stakeholders_orgs_list, start_year, end_year, quota_dates)
    stakeholders_pers_sum_all_persons, stakeholders_orgs_sum_all_orgs = results[9], results[10]
    
    # Prepare the final table for organizations with the desired format
    orgs_table = common_orgs[['Name', 'Stakeholder-Gruppe', 'Governance level', 'Value']].fillna('')
    orgs_table = orgs_table.sort_values(by=['Stakeholder-Gruppe', 'Governance level', 'Value'], ascending=[True, True, False])
    
    # Prepare the final table for persons with the desired format
    pers_table = common_pers[['Name', 'Stakeholder-Gruppe', 'Governance level', 'Value']].fillna('')
    pers_table = pers_table.sort_values(by=['Stakeholder-Gruppe', 'Governance level', 'Value'], ascending=[True, True, False])
    
    # Create a summary table for organizations
    orgs_summary = orgs_table.groupby(['Stakeholder-Gruppe', 'Governance level']).sum().reset_index()
    orgs_summary['Type'] = 'Organizations'
    
    # Create a summary table for persons
    pers_summary = pers_table.groupby(['Stakeholder-Gruppe', 'Governance level']).sum().reset_index()
    pers_summary['Type'] = 'Individuals'
    
    summary_table = pd.concat([orgs_summary, pers_summary])
    
    # Compute subtotals by stakeholder group
    subtotals = summary_table.groupby(['Stakeholder-Gruppe', 'Type']).sum().reset_index()
    
    # Compute total for each type across all stakeholder groups
    total = summary_table.groupby('Type').sum().reset_index()
    total['Stakeholder-Gruppe'] = 'Total'
    
    # Compute grand total for each stakeholder group across all types
    grand_totals = summary_table.groupby('Stakeholder-Gruppe').sum().reset_index()
    grand_totals['Type'] = 'All Types'
    grand_totals['Governance level'] = 'Grand Total'
    
    # Adding subtotals and sorting
    summary_table_with_subtotals = pd.DataFrame()
    for stakeholder_group in summary_table['Stakeholder-Gruppe'].unique():
        for type_ in summary_table[summary_table['Stakeholder-Gruppe'] == stakeholder_group]['Type'].unique():
            filtered_data = summary_table[(summary_table['Stakeholder-Gruppe'] == stakeholder_group) & (summary_table['Type'] == type_)]
            summary_table_with_subtotals = pd.concat([summary_table_with_subtotals, filtered_data])
    
            # Add subtotal row for current group and type
            subtotal_row = subtotals[(subtotals['Stakeholder-Gruppe'] == stakeholder_group) & (subtotals['Type'] == type_)]
            summary_table_with_subtotals = pd.concat([summary_table_with_subtotals, subtotal_row])
    
        # Add grand total row after processing all types for a stakeholder group
        grand_total_row = grand_totals[grand_totals['Stakeholder-Gruppe'] == stakeholder_group]
        summary_table_with_subtotals = pd.concat([summary_table_with_subtotals, grand_total_row])
    
    # Append total rows at the end
    summary_table_with_subtotals = pd.concat([summary_table_with_subtotals, total]) 
    summary_table_with_subtotals['Governance level'].fillna('All Gov Levels', inplace=True)
    
    # Output path for the Excel file
    output_path = 'Results\stakeholder_analysis_total_count_tab4c.xlsx'
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        workbook = writer.book
    
        # Formatting
        bold = workbook.add_format({'bold': True})
        italic = workbook.add_format({'italic': True})
    
        worksheet = workbook.add_worksheet('Summary')
        worksheet.write('A1', 'Stakeholder Group', bold)
        worksheet.write('B1', 'Governance Level', bold)
        worksheet.write('C1', 'Type', bold)
        worksheet.write('D1', 'Value', bold)
    
        start_row = 1
        for stakeholder_group in summary_table_with_subtotals['Stakeholder-Gruppe'].unique():
            worksheet.write(start_row, 0, stakeholder_group, italic)
            start_row += 1
    
            for type_ in summary_table_with_subtotals[summary_table_with_subtotals['Stakeholder-Gruppe'] == stakeholder_group]['Type'].unique():
                worksheet.write(start_row, 1, type_, italic)
                start_row += 1
                
                for _, row in summary_table_with_subtotals[(summary_table_with_subtotals['Stakeholder-Gruppe'] == stakeholder_group) & (summary_table_with_subtotals['Type'] == type_)].iterrows():
                    worksheet.write(start_row, 0, row['Stakeholder-Gruppe'])
                    worksheet.write(start_row, 1, row['Governance level'])
                    worksheet.write(start_row, 2, row['Type'])
                    worksheet.write(start_row, 3, row['Value'])
                    start_row += 1
                start_row += 1
    
            start_row += 1

    stakeholders_pers_sum_all_persons.reset_index(inplace=True)
    stakeholders_pers_sum_all_persons['Name'] = stakeholders_pers_sum_all_persons['Name'].str.lower().str.strip()
    common_pers['Name'] = common_pers['Name'].str.lower().str.strip()

    common_pers = pd.merge(common_pers, stakeholders_pers_sum_all_persons, on='Name', how='inner')

    stakeholders_orgs_sum_all_orgs.reset_index(inplace=True)
    stakeholders_orgs_sum_all_orgs['Name'] = stakeholders_orgs_sum_all_orgs['Name'].str.lower().str.strip()
    common_orgs['Name'] = common_orgs['Name'].str.lower().str.strip()

    common_orgs = pd.merge(common_orgs, stakeholders_orgs_sum_all_orgs, on='Name', how='inner')

    stakeholders_pers_list = list(common_pers[common_pers['Stakeholder-Gruppe'].str.contains('politics|management|science|fisheries|engo', na=False)]['Name'])
    stakeholders_orgs_list = list(common_orgs[common_orgs['Stakeholder-Gruppe'].str.contains('politics|management|science|fisheries|engo', na=False)]['Name'])

    # Run stakeholder analysis again with filtered data
    results = run_stakeholder_analysis(data, common_pers, common_orgs, stakeholders_pers_list, stakeholders_orgs_list, start_year, end_year, quota_dates)
    
    # Further analysis based on stakeholder groups
    groups = ['politics|management', 'fisheries', 'science', 'engo']
    analysis_results = {group: analyze_by_group(common_pers, common_orgs, group, data, start_year, end_year, quota_dates) for group in groups}
    
    # Summary statistics for each group and governance level
    summary_stats = {}
    
    summary_stats['all_persons_sentiment'] = calculate_summary_stats(results[1]['Sentiment Index'])
    summary_stats['all_organizations_sentiment'] = calculate_summary_stats(results[3]['Sentiment Index'])
    summary_stats['all_both_sentiment'] = calculate_summary_stats(results[5]['Sentiment Index'])
    
    summary_stats['all_persons_share'] = calculate_summary_stats(results[0]['share'])
    summary_stats['all_organizations_share'] = calculate_summary_stats(results[2]['share'])
    summary_stats['all_both_share'] = calculate_summary_stats(results[4]['share'])
    
    summary_stats['all_persons_share_once'] = calculate_summary_stats(results[0]['share_once'])
    summary_stats['all_organizations_share_once'] = calculate_summary_stats(results[2]['share_once'])
    summary_stats['all_both_share_once'] = calculate_summary_stats(results[4]['share_once'])

    for group, result in analysis_results.items():
        stakeholders_pers_sum = result[0]
        sentiment_pers_sum = result[1]
        
        stakeholders_orgs_sum = result[2]
        sentiment_orgs_sum = result[3]
        
        stakeholders_both_sum = result[4]
        sentiment_both_sum = result[5]
        
        summary_stats[f'{group}_persons_sentiment'] = calculate_summary_stats(sentiment_pers_sum['Sentiment Index'])
        summary_stats[f'{group}_organizations_sentiment'] = calculate_summary_stats(sentiment_orgs_sum['Sentiment Index'])
        summary_stats[f'{group}_both_sentiment'] = calculate_summary_stats(sentiment_both_sum['Sentiment Index'])
        
        summary_stats[f'{group}_persons_sentiment_pos'] = calculate_summary_stats(sentiment_pers_sum['pos_sent_ratio'])
        summary_stats[f'{group}_organizations_sentiment_pos'] = calculate_summary_stats(sentiment_orgs_sum['pos_sent_ratio'])
        summary_stats[f'{group}_both_sentiment_pos'] = calculate_summary_stats(sentiment_both_sum['pos_sent_ratio'])
        
        summary_stats[f'{group}_persons_sentiment_neu'] = calculate_summary_stats(sentiment_pers_sum['neu_sent_ratio'])
        summary_stats[f'{group}_organizations_sentiment_neu'] = calculate_summary_stats(sentiment_orgs_sum['neu_sent_ratio'])
        summary_stats[f'{group}_both_sentiment_neu'] = calculate_summary_stats(sentiment_both_sum['neu_sent_ratio'])
        
        summary_stats[f'{group}_persons_sentiment_neg'] = calculate_summary_stats(sentiment_pers_sum['neg_sent_ratio'])
        summary_stats[f'{group}_organizations_sentiment_neg'] = calculate_summary_stats(sentiment_orgs_sum['neg_sent_ratio'])
        summary_stats[f'{group}_both_sentiment_neg'] = calculate_summary_stats(sentiment_both_sum['neg_sent_ratio'])
        
        summary_stats[f'{group}_persons_share'] = calculate_summary_stats(stakeholders_pers_sum['share'])
        summary_stats[f'{group}_organizations_share'] = calculate_summary_stats(stakeholders_orgs_sum['share'])
        summary_stats[f'{group}_both_share'] = calculate_summary_stats(stakeholders_both_sum['share'])
        
        summary_stats[f'{group}_persons_share_once'] = calculate_summary_stats(stakeholders_pers_sum['share_once'])
        summary_stats[f'{group}_organizations_share_once'] = calculate_summary_stats(stakeholders_orgs_sum['share_once'])
        summary_stats[f'{group}_both_share_once'] = calculate_summary_stats(stakeholders_both_sum['share_once'])
    
    for group, result in analysis_results.items():
        stakeholders_both_sum = result[4]
        stakeholders_both_sum.index = pd.to_datetime(stakeholders_both_sum.index, format='%Y')
        
    write_summary_to_excel(summary_stats,'Results\Summary_Statistics_tabS6_S7_S8_S9_S10.xlsx')
    create_summary_excel_2(summary_stats,'Results\Stakeholder_counts_tab1_tabS9.xlsx')

    return results, analysis_results, quota_dates, summary_stats

if __name__ == "__main__":
    results, analysis_results, quota_dates, summary_stats = main(data_path, pers_path, orgs_path, quota_path, start_year, end_year)
        
    stakeholders_both_sum_pm = analysis_results['politics|management'][4]
    stakeholders_both_sum_f = analysis_results['fisheries'][4]
    stakeholders_both_sum_s = analysis_results['science'][4]
    stakeholders_both_sum_e = analysis_results['engo'][4]

    stakeholders_sentiment_both_pm = analysis_results['politics|management'][5].reset_index(drop=True)
    stakeholders_sentiment_both_f = analysis_results['fisheries'][5].reset_index(drop=True)
    stakeholders_sentiment_both_s = analysis_results['science'][5].reset_index(drop=True)
    stakeholders_sentiment_both_e = analysis_results['engo'][5].reset_index(drop=True)
    
    stakeholders_pers_sum_pm = analysis_results['politics|management'][0]
    stakeholders_pers_sum_f = analysis_results['fisheries'][0]
    stakeholders_pers_sum_s = analysis_results['science'][0]
    stakeholders_pers_sum_e = analysis_results['engo'][0]

    stakeholders_sentiment_pers_pm = analysis_results['politics|management'][1].reset_index(drop=True)
    stakeholders_sentiment_pers_f = analysis_results['fisheries'][1].reset_index(drop=True)
    stakeholders_sentiment_pers_s = analysis_results['science'][1].reset_index(drop=True)
    stakeholders_sentiment_pers_e = analysis_results['engo'][1].reset_index(drop=True)
    
    stakeholders_orgs_sum_pm = analysis_results['politics|management'][2]
    stakeholders_orgs_sum_f = analysis_results['fisheries'][2]
    stakeholders_orgs_sum_s = analysis_results['science'][2]
    stakeholders_orgs_sum_e = analysis_results['engo'][2]

    stakeholders_sentiment_orgs_pm = analysis_results['politics|management'][3].reset_index(drop=True)
    stakeholders_sentiment_orgs_f = analysis_results['fisheries'][3].reset_index(drop=True)
    stakeholders_sentiment_orgs_s = analysis_results['science'][3].reset_index(drop=True)
    stakeholders_sentiment_orgs_e = analysis_results['engo'][3].reset_index(drop=True)
    
    years = quota_dates[:-1].dt.year

    # Save results
    df_both_sum_pm = pd.DataFrame({'Years': years, 
                                   'Stakeholder Share (multiple in article)': stakeholders_both_sum_pm['share'].values,
                                   'Stakeholder Share (only once in article)': stakeholders_both_sum_pm['share_once'].values})
    df_both_sum_f = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_both_sum_f['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_both_sum_f['share_once'].values})
    df_both_sum_s = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_both_sum_s['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_both_sum_s['share_once'].values})
    df_both_sum_e = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_both_sum_e['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_both_sum_e['share_once'].values})

    df_sentiment_both_pm = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_both_pm.drop(columns=['year'])], axis=1)
    df_sentiment_both_f = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_both_f.drop(columns=['year'])], axis=1)
    df_sentiment_both_s = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_both_s.drop(columns=['year'])], axis=1)
    df_sentiment_both_e = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_both_e.drop(columns=['year'])], axis=1)
    
    df_sentiment_pers_pm = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_pers_pm.drop(columns=['year'])], axis=1)
    df_sentiment_pers_f = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_pers_f.drop(columns=['year'])], axis=1)
    df_sentiment_pers_s = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_pers_s.drop(columns=['year'])], axis=1)
    df_sentiment_pers_e = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_pers_e.drop(columns=['year'])], axis=1)
    
    df_sentiment_orgs_pm = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_orgs_pm.drop(columns=['year'])], axis=1)
    df_sentiment_orgs_f = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_orgs_f.drop(columns=['year'])], axis=1)
    df_sentiment_orgs_s = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_orgs_s.drop(columns=['year'])], axis=1)
    df_sentiment_orgs_e = pd.concat([pd.DataFrame({'Fishery Year': years}), stakeholders_sentiment_orgs_e.drop(columns=['year'])], axis=1)
    
    df_pers_sum_pm = pd.DataFrame({'Years': years, 
                                   'Stakeholder Share (multiple in article)': stakeholders_pers_sum_pm['share'].values,
                                   'Stakeholder Share (only once in article)': stakeholders_pers_sum_pm['share_once'].values})
    df_pers_sum_f = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_pers_sum_f['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_pers_sum_f['share_once'].values})
    df_pers_sum_s = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_pers_sum_s['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_pers_sum_s['share_once'].values})
    df_pers_sum_e = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_pers_sum_e['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_pers_sum_e['share_once'].values})
    
    df_orgs_sum_pm = pd.DataFrame({'Years': years, 
                                   'Stakeholder Share (multiple in article)': stakeholders_orgs_sum_pm['share'].values,
                                   'Stakeholder Share (only once in article)': stakeholders_orgs_sum_pm['share_once'].values})
    df_orgs_sum_f = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_orgs_sum_f['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_orgs_sum_f['share_once'].values})
    df_orgs_sum_s = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_orgs_sum_s['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_orgs_sum_s['share_once'].values})
    df_orgs_sum_e = pd.DataFrame({'Years': years, 
                                  'Stakeholder Share (multiple in article)': stakeholders_orgs_sum_e['share'].values,
                                  'Stakeholder Share (only once in article)': stakeholders_orgs_sum_e['share_once'].values})
    
    # Writing to Excel
    output_path = 'Results\stakeholder_analysis_time_series.xlsx'
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:

        df_both_sum_pm.to_excel(writer, sheet_name='Both_PM_Summary', index=False)
        df_both_sum_f.to_excel(writer, sheet_name='Both_F_Summary', index=False)
        df_both_sum_s.to_excel(writer, sheet_name='Both_S_Summary', index=False)
        df_both_sum_e.to_excel(writer, sheet_name='Both_E_Summary', index=False)
        
        df_sentiment_both_pm.to_excel(writer, sheet_name='Both_PM_Sentiment', index=False)
        df_sentiment_both_f.to_excel(writer, sheet_name='Both_F_Sentiment', index=False)
        df_sentiment_both_s.to_excel(writer, sheet_name='Both_S_Sentiment', index=False)
        df_sentiment_both_e.to_excel(writer, sheet_name='Both_E_Sentiment', index=False)
    
        df_pers_sum_pm.to_excel(writer, sheet_name='Pers_PM_Summary', index=False)
        df_pers_sum_f.to_excel(writer, sheet_name='Pers_F_Summary', index=False)
        df_pers_sum_s.to_excel(writer, sheet_name='Pers_S_Summary', index=False)
        df_pers_sum_e.to_excel(writer, sheet_name='Pers_E_Summary', index=False)
        
        df_sentiment_pers_pm.to_excel(writer, sheet_name='Pers_PM_Sentiment', index=False)
        df_sentiment_pers_f.to_excel(writer, sheet_name='Pers_F_Sentiment', index=False)
        df_sentiment_pers_s.to_excel(writer, sheet_name='Pers_S_Sentiment', index=False)
        df_sentiment_pers_e.to_excel(writer, sheet_name='Pers_E_Sentiment', index=False)
    
        df_orgs_sum_pm.to_excel(writer, sheet_name='Orgs_PM_Summary', index=False)
        df_orgs_sum_f.to_excel(writer, sheet_name='Orgs_F_Summary', index=False)
        df_orgs_sum_s.to_excel(writer, sheet_name='Orgs_S_Summary', index=False)
        df_orgs_sum_e.to_excel(writer, sheet_name='Orgs_E_Summary', index=False)
        
        df_sentiment_orgs_pm.to_excel(writer, sheet_name='Orgs_PM_Sentiment', index=False)
        df_sentiment_orgs_f.to_excel(writer, sheet_name='Orgs_F_Sentiment', index=False)
        df_sentiment_orgs_s.to_excel(writer, sheet_name='Orgs_S_Sentiment', index=False)
        df_sentiment_orgs_e.to_excel(writer, sheet_name='Orgs_E_Sentiment', index=False)
   
    ###########################################################################
    
    # Generate figure 5
    fig, ax1 = plt.subplots(dpi=300)
    
    color1 = (0, 89, 84) 
    color2 = (254, 217, 145)
    color3 = (99, 194, 203) 
    color4 = (244,177,131) 
    
    # Convert RGB to matplotlib color format
    def convert_color(rgb):
        return tuple([x / 255. for x in rgb])
    
    color1 = convert_color(color1)
    color2 = convert_color(color2)
    color3 = convert_color(color3)
    color4 = convert_color(color4)
    
    wi = 200
    
    ax1.plot(quota_dates[:-1], np.array(stakeholders_sentiment_both_pm['Sentiment Index']), color=color1, linewidth=3, linestyle='-', label='Politics & Public Authorities')
    ax1.plot(quota_dates[:-1], np.array(stakeholders_sentiment_both_f['Sentiment Index']), color=color2, linewidth=3, linestyle='-', label='Fishery')
    ax1.plot(quota_dates[:-1], np.array(stakeholders_sentiment_both_s['Sentiment Index']), color=color3, linewidth=3, linestyle='-', label='Science')
    ax1.plot(quota_dates[:-1], np.array(stakeholders_sentiment_both_e['Sentiment Index']), color=color4, linewidth=3, linestyle='-', label='eNGO') 
    
    ax1.set_xlim(datetime.datetime(2009, 1, 1), datetime.datetime(2020, 12, 31))
    ax1.tick_params(axis='y', labelsize = 18, which='both', left=True, labelleft=True)
    ax1.tick_params(axis='x', labelsize = 18)
    ax1.set_ylabel('Sentiment', fontsize = 24)
    ax1.set_xticks(ax1.get_xticks()[1:-1])
    ax1.set_xticklabels(ax1.get_xticklabels(), rotation=45)
    ax1.xaxis.set_major_locator(mdates.YearLocator())
    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    ax1.set_ylim(-0.6, -0.05)
    
    fig.subplots_adjust(left=0.15, right=0.875, top=0.95, bottom=0.15)

    handles, labels = ax1.get_legend_handles_labels()
    fig.legend(handles, labels, loc='lower center', bbox_to_anchor=(0.5, -0.12), shadow=False, ncol=4, fontsize=16, frameon=False)
    
    plt.savefig(r'Figures\figure5.svg', format='svg', bbox_inches='tight')
    plt.savefig(r'Figures\figure5.eps', format='eps', bbox_inches='tight')
    plt.show()